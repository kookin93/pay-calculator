from pathlib import Path
from decimal import Decimal, ROUND_HALF_UP, ROUND_FLOOR
import bisect

import streamlit as st

try:
    import openpyxl
except Exception:
    openpyxl = None


APP_DIR = Path(__file__).resolve().parent

LOGO_FILENAME = "로고_검정색.png"

# GitHub 저장소 최상단에 이 이름으로 업로드해야 합니다
TAX_XLSX_FILENAME = "tax_table.xlsx"

OFFICE_NAME = "인화세무회계컨설팅"
OFFICE_PHONE = "042-222-7208"
OFFICE_ADDRESS = "대전 중구 충무로 173 대현빌딩 6층"

WEEKS_PER_MONTH = Decimal("4.345")

# 엑셀 AI19 AJ19 AK19 AL19
PENSION_RATE = Decimal("0.045")
HEALTH_RATE = Decimal("0.03545")
CARE_RATE = Decimal("0.1295")
EMPLOY_RATE = Decimal("0.009")


def d(x) -> Decimal:
    return Decimal(str(x))


def round0(x: Decimal) -> Decimal:
    return x.quantize(Decimal("1"), rounding=ROUND_HALF_UP)


def floor_to_step(x: Decimal, step: Decimal) -> Decimal:
    if step == 0:
        return x
    return (x / step).to_integral_value(rounding=ROUND_FLOOR) * step


def won(x: Decimal) -> str:
    return f"{int(x):,}원"


def find_tax_xlsx_path() -> Path:
    p = APP_DIR / TAX_XLSX_FILENAME
    return p


@st.cache_data(show_spinner=False)
def load_withholding_table() -> dict:
    """
    엑셀
    간이세액표 A5:M651
    - A 이상
    - B 미만
    - C~M 공제대상가족수 1~11
    """
    if openpyxl is None:
        return {"ok": False, "message": "openpyxl 미설치입니다. requirements.txt에 openpyxl을 추가하세요", "rows": [], "files": []}

    p = find_tax_xlsx_path()
    files = sorted([x.name for x in APP_DIR.iterdir() if x.is_file()])

    if not p.exists():
        return {"ok": False, "message": f"엑셀 파일을 찾지 못했습니다. 파일명은 {TAX_XLSX_FILENAME} 입니다", "rows": [], "files": files}

    wb = openpyxl.load_workbook(p, data_only=True)
    if "간이세액표" not in wb.sheetnames:
        return {"ok": False, "message": "엑셀에 간이세액표 시트가 없습니다. 시트명이 정확히 간이세액표 여야 합니다", "rows": [], "files": files}

    ws = wb["간이세액표"]

    rows = []
    for r in range(6, 652):  # 6~651
        low = ws.cell(r, 1).value
        high = ws.cell(r, 2).value

        if low is None or high is None:
            continue
        if not isinstance(low, (int, float)) or not isinstance(high, (int, float)):
            continue

        vals = []
        for c in range(3, 14):  # C~M
            v = ws.cell(r, c).value
            if v is None:
                v = 0
            vals.append(int(v))

        rows.append((int(low), int(high), vals))

    if not rows:
        return {"ok": False, "message": "간이세액표 데이터가 비어 있습니다", "rows": [], "files": files}

    rows.sort(key=lambda x: x[0])
    return {"ok": True, "message": "ok", "rows": rows, "files": files}


def vlookup_approx(monthly_taxable: int, family_count: int, table_rows: list) -> int:
    """
    엑셀
    VLOOKUP(AF20, 간이세액표!A5:M651, (2+AR20), 1)
    family_count는 1~11
    """
    if family_count < 1:
        family_count = 1
    if family_count > 11:
        family_count = 11

    lows = [r[0] for r in table_rows]
    idx = bisect.bisect_right(lows, monthly_taxable) - 1
    if idx < 0:
        raise KeyError("table underflow")

    low, high, vals = table_rows[idx]
    return int(vals[family_count - 1])


def child_tax_credit(children_8to20: int) -> int:
    """
    엑셀 AG20 마지막 차감부
    IF(AS20=0,0,IF(AS20=1,12500,IF(AS20=2,29160,29160+(AS20-2)*25000)))
    """
    if children_8to20 <= 0:
        return 0
    if children_8to20 == 1:
        return 12500
    if children_8to20 == 2:
        return 29160
    return 29160 + (children_8to20 - 2) * 25000


def compute_income_tax_from_excel_logic(
    monthly_taxable: int,
    family_count: int,
    children_8to20: int,
) -> Decimal:
    """
    엑셀 AG20 구현
    =IFERROR(ROUNDDOWN(구간별 계산 또는 VLOOKUP, -1), 0) - 자녀세액공제
    """
    tbl = load_withholding_table()
    base_tax = 0

    try:
        AF = monthly_taxable

        if AF > 87000000:
            v = Decimal("1507400") + Decimal("31034600") + d(AF - 87000000) * Decimal("0.45")
        elif AF > 45000000:
            v = Decimal("1507400") + Decimal("13394600") + d(AF - 45000000) * Decimal("0.42")
        elif AF > 30000000:
            v = Decimal("1507400") + Decimal("7394600") + d(AF - 30000000) * Decimal("0.4")
        elif AF > 28000000:
            v = Decimal("1507400") + Decimal("6610600") + d(AF - 28000000) * Decimal("0.98") * Decimal("0.4")
        elif AF > 14000000:
            v = Decimal("1507400") + Decimal("1397000") + d(AF - 14000000) * Decimal("0.98") * Decimal("0.38")
        elif AF > 10000000:
            v = Decimal("1507400") + Decimal("25000") + d(AF - 10000000) * Decimal("0.98") * Decimal("0.35")
        elif AF == 10000000:
            v = Decimal("1507400")
        else:
            if not tbl["ok"]:
                raise KeyError(tbl["message"])
            vlookup_val = vlookup_approx(AF, family_count, tbl["rows"])
            v = d(vlookup_val)

        base_tax = int(floor_to_step(v, d("10")))  # ROUNDDOWN( , -1)

    except Exception:
        base_tax = 0

    credit = child_tax_credit(children_8to20)
    return d(base_tax - credit)


def compute(
    annual_salary: Decimal,
    daily_work_hours: Decimal,
    weekly_ot_hours: Decimal,
    work_days_per_week: Decimal,
    min_wage: Decimal,
    min_inclusion_ratio: Decimal,
    meal: Decimal,
    car: Decimal,
    child: Decimal,
    duty: Decimal,
    grade: Decimal,
    etc_allow: Decimal,
    etc_deduct: Decimal,
    family_count: int,
    children_8to20: int,
) -> dict:
    # AR7 = ROUND(AR6*209,0)
    min_monthly_wage = round0(min_wage * d("209"))

    # AR11 = ROUND(AR7*AR10,0)
    min_non_included = round0(min_monthly_wage * min_inclusion_ratio)

    # Y = ROUND(Z/12,0)
    monthly_pay = round0(annual_salary / d("12"))

    # M = ROUND(((L*근로일)+L)*4.345,0)
    monthly_standard_hours = round0(((daily_work_hours * work_days_per_week) + daily_work_hours) * WEEKS_PER_MONTH)

    # Q = P*4.345
    monthly_ot_hours = weekly_ot_hours * WEEKS_PER_MONTH

    # R = M + (P*1.5*4.345)
    weighted_total_hours = monthly_standard_hours + (weekly_ot_hours * d("1.5") * WEEKS_PER_MONTH)
    if weighted_total_hours == 0:
        raise ValueError("시간 값이 0이라 계산할 수 없습니다")

    # AB = Y / R
    hourly = monthly_pay / weighted_total_hours

    # O = ROUND(Q*AB*1.5,0)
    fixed_ot_pay = round0(monthly_ot_hours * hourly * d("1.5"))

    # S~X
    allowances_sum = meal + car + child + duty + grade + etc_allow

    # N = Y - (O + S~X)
    base_pay = monthly_pay - (fixed_ot_pay + allowances_sum)

    check_ok = (monthly_pay == (base_pay + fixed_ot_pay + allowances_sum))

    # AC
    non_tax_sum = meal + car + child
    if non_tax_sum > 0:
        compare_wage = (base_pay + non_tax_sum - min_non_included) / monthly_standard_hours
    else:
        compare_wage = (base_pay + non_tax_sum) / monthly_standard_hours

    compliance = "준수" if compare_wage >= min_wage else "미준수"

    # AF
    taxable = monthly_pay - non_tax_sum

    # AG
    income_tax = compute_income_tax_from_excel_logic(
        monthly_taxable=int(taxable),
        family_count=family_count,
        children_8to20=children_8to20,
    )

    # AH
    resident_tax = d(floor_to_step(income_tax * d("0.1"), d("10")))

    # AI
    pension_base = floor_to_step(taxable, d("1000"))
    pension = d(floor_to_step(pension_base * PENSION_RATE, d("10")))

    # AJ AK AL
    health = d(floor_to_step(taxable * HEALTH_RATE, d("10")))
    care = d(floor_to_step(health * CARE_RATE, d("10")))
    employ = d(floor_to_step(taxable * EMPLOY_RATE, d("10")))

    # AN
    total_deduct = income_tax + resident_tax + pension + health + care + employ + etc_deduct

    # AO
    net_pay = monthly_pay - total_deduct

    diff = (base_pay + fixed_ot_pay + allowances_sum) - monthly_pay
    diff_abs = abs(diff)

    return {
        "monthly_pay": monthly_pay,
        "monthly_standard_hours": monthly_standard_hours,
        "hourly": round0(hourly),
        "base_pay": round0(base_pay),
        "fixed_ot_pay": fixed_ot_pay,
        "taxable": round0(taxable),
        "income_tax": round0(income_tax),
        "resident_tax": round0(resident_tax),
        "pension": round0(pension),
        "health": round0(health),
        "care": round0(care),
        "employ": round0(employ),
        "etc_deduct": round0(etc_deduct),
        "total_deduct": round0(total_deduct),
        "net_pay": round0(net_pay),
        "compare_wage": round0(compare_wage),
        "min_wage": round0(min_wage),
        "compliance": compliance,
        "diff_abs": round0(diff_abs),
        "check_ok": check_ok,
    }


def render_footer() -> None:
    st.markdown(
        f"""
        <style>
          .footer {{
            position: fixed;
            left: 0;
            right: 0;
            bottom: 0;
            background: #ffffff;
            border-top: 1px solid #e6e6e6;
            padding: 10px 12px;
            font-size: 13px;
            color: #444;
            z-index: 9999;
          }}
          .footer-inner {{
            max-width: 860px;
            margin: 0 auto;
            display: flex;
            justify-content: space-between;
            gap: 12px;
            flex-wrap: wrap;
          }}
          .footer strong {{
            font-weight: 700;
          }}
          .wrap-pad {{
            padding-bottom: 78px;
          }}
        </style>
        <div class="footer">
          <div class="footer-inner">
            <div><strong>{OFFICE_NAME}</strong></div>
            <div>{OFFICE_PHONE}</div>
            <div>{OFFICE_ADDRESS}</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


st.set_page_config(page_title="급여 계산기", layout="centered")

logo_path = APP_DIR / LOGO_FILENAME
if logo_path.exists():
    st.image(str(logo_path), use_container_width=True)

st.markdown("## 급여 계산기")

st.warning(
    "저작권 안내\n"
    "본 프로그램과 화면 구성 및 산출물은 저작권 보호 대상입니다\n"
    "사전 서면 동의 없이 복제 수정 배포 게시를 금지합니다"
)

st.info(
    "참고 안내\n"
    "본 계산 결과는 참고용입니다\n"
    "법적 자문 또는 확정 금액이 아닙니다\n"
    "최종 판단과 책임은 사용자에게 있습니다"
)

tbl_status = load_withholding_table()
if not tbl_status["ok"]:
    st.error(f"간이세액표 로드 실패. {tbl_status['message']}")
    if "files" in tbl_status and tbl_status["files"]:
        st.caption("현재 GitHub 폴더에 보이는 파일 목록입니다")
        st.code("\n".join(tbl_status["files"]))
    st.caption(f"엑셀 파일명은 {TAX_XLSX_FILENAME} 여야 합니다")
else:
    st.caption("소득세는 엑셀 간이세액표 시트를 그대로 사용합니다")

with st.sidebar:
    st.header("입력")

    annual_salary_int = st.number_input("상여금 제외 연봉", min_value=0, value=36000000, step=100000, format="%d")

    work_days = st.number_input("주 근로일", min_value=1, value=5, step=1, format="%d")
    daily_hours = st.number_input("1일 근로시간", min_value=0.0, value=8.0, step=0.5)

    weekly_ot = st.number_input("주 고정연장시간", min_value=0.0, value=9.0, step=0.5)

    st.divider()
    st.subheader("최저임금")
    min_wage_int = st.number_input("최저시급", min_value=0, value=10320, step=10, format="%d")
    inclusion_ratio = st.number_input("최저 산입비율", min_value=0.0, max_value=1.0, value=0.0, step=0.01)

    st.divider()
    st.subheader("소득세 입력")
    family_count = st.number_input("공제대상가족수", min_value=1, max_value=11, value=1, step=1, format="%d")
    children_8to20 = st.number_input("8세 이상 20세 이하", min_value=0, value=0, step=1, format="%d")

    st.divider()
    st.subheader("비과세 수당")
    meal_int = st.number_input("식대", min_value=0, value=0, step=10000, format="%d")
    car_int = st.number_input("차량유지비", min_value=0, value=0, step=10000, format="%d")
    child_int = st.number_input("자녀 양육수당", min_value=0, value=0, step=10000, format="%d")

    st.divider()
    st.subheader("과세 수당")
    duty_int = st.number_input("직책수당", min_value=0, value=0, step=10000, format="%d")
    grade_int = st.number_input("직급수당", min_value=0, value=0, step=10000, format="%d")
    etc_allow_int = st.number_input("기타수당", min_value=0, value=0, step=10000, format="%d")

    st.divider()
    st.subheader("기타 공제")
    etc_deduct_int = st.number_input("기타공제", min_value=0, value=0, step=10000, format="%d")

try:
    r = compute(
        annual_salary=d(annual_salary_int),
        daily_work_hours=d(daily_hours),
        weekly_ot_hours=d(weekly_ot),
        work_days_per_week=d(work_days),
        min_wage=d(min_wage_int),
        min_inclusion_ratio=d(inclusion_ratio),
        meal=d(meal_int),
        car=d(car_int),
        child=d(child_int),
        duty=d(duty_int),
        grade=d(grade_int),
        etc_allow=d(etc_allow_int),
        etc_deduct=d(etc_deduct_int),
        family_count=int(family_count),
        children_8to20=int(children_8to20),
    )
except Exception as e:
    st.error(str(e))
    st.stop()

st.markdown('<div class="wrap-pad">', unsafe_allow_html=True)

st.subheader("계산 결과")
col1, col2 = st.columns(2)
with col1:
    st.metric("총 월급여액", won(r["monthly_pay"]))
    st.metric("기본급", won(r["base_pay"]))
    st.metric("고정연장수당", won(r["fixed_ot_pay"]))
with col2:
    st.metric("통상시급", won(r["hourly"]))
    st.metric("과세금액", won(r["taxable"]))
    st.metric("실지급액", won(r["net_pay"]))

st.divider()
st.subheader("최저임금")
c1, c2, c3 = st.columns(3)
with c1:
    st.metric("최저시급", won(r["min_wage"]))
with c2:
    st.metric("비교기준임금", won(r["compare_wage"]))
with c3:
    st.metric("준수여부", r["compliance"])

st.divider()
st.subheader("공제")
c1, c2 = st.columns(2)
with c1:
    st.write("소득세")
    st.write(won(r["income_tax"]))
    st.write("주민세")
    st.write(won(r["resident_tax"]))
    st.write("국민연금")
    st.write(won(r["pension"]))
    st.write("건강보험")
    st.write(won(r["health"]))
with c2:
    st.write("장기요양보험")
    st.write(won(r["care"]))
    st.write("고용보험")
    st.write(won(r["employ"]))
    st.write("기타공제")
    st.write(won(r["etc_deduct"]))
    st.metric("총 공제금액", won(r["total_deduct"]))

st.divider()
st.subheader("검증")
if r["diff_abs"] <= Decimal("1"):
    st.success(f"정상. 차이 절대값 {int(r['diff_abs']):,}원")
else:
    st.error(f"비정상. 차이 절대값 {int(r['diff_abs']):,}원")

if r["check_ok"]:
    st.caption("CHECK. OK")
else:
    st.caption("CHECK. NOTOK")

render_footer()
st.markdown("</div>", unsafe_allow_html=True)
